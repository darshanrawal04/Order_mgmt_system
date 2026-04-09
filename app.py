from flask import Flask, request, jsonify, send_file, session, Response
from functools import wraps
import json, os, hashlib, secrets, re
from datetime import datetime, timedelta, timezone
from urllib.parse import quote

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", secrets.token_hex(32))
app.permanent_session_lifetime = timedelta(days=30)

# ── Paths ─────────────────────────────────────────────────────────────────────
DATA_DIR       = "data"
ORDERS_FILE    = f"{DATA_DIR}/orders.json"
CUSTOMERS_FILE = f"{DATA_DIR}/customers.json"
ADMIN_FILE     = f"{DATA_DIR}/admin.json"
PUSH_SUBS_FILE = f"{DATA_DIR}/push_subscriptions.json"
NOTES_FILE     = f"{DATA_DIR}/customer_notes.json"
ROUTES_FILE    = f"{DATA_DIR}/routes.json"
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs("templates", exist_ok=True)

# ── IST ───────────────────────────────────────────────────────────────────────
IST = timezone(timedelta(hours=5, minutes=30))
ORDER_CUTOFF_HOUR   = 14
ORDER_CUTOFF_MINUTE = 30

def now_ist():
    return datetime.now(IST)

def is_order_open():
    n = now_ist()
    cutoff = n.replace(hour=ORDER_CUTOFF_HOUR, minute=ORDER_CUTOFF_MINUTE, second=0, microsecond=0)
    return n < cutoff

def cutoff_str(): return "2:30 PM"

# ── Products ──────────────────────────────────────────────────────────────────
# UPDATED: renamed buttermilk, added Banas, added full Curd category
PRODUCTS = [
    # Milk
    {"id":"amul_gold_500ml",        "name":"Amul Gold",             "variant":"500 mL",  "category":"Milk",       "emoji":"🥛"},
    {"id":"amul_gold_6ltr",         "name":"Amul Gold",             "variant":"6 Ltr",   "category":"Milk",       "emoji":"🥛"},
    {"id":"amul_taaza_500ml",       "name":"Amul Taaza",            "variant":"500 mL",  "category":"Milk",       "emoji":"🍼"},
    {"id":"amul_taaza_6ltr",        "name":"Amul Taaza",            "variant":"6 Ltr",   "category":"Milk",       "emoji":"🍼"},
    {"id":"amul_taaza_pouch",       "name":"Amul Taaza",            "variant":"Pouch",   "category":"Milk",       "emoji":"🍼"},
    # Buttermilk (renamed + added Banas)
    {"id":"buttermilk_500ml",       "name":"Amul Buttermilk",       "variant":"500 mL",  "category":"Buttermilk", "emoji":"🥤"},
    {"id":"buttermilk_6ltr",        "name":"Amul Buttermilk",       "variant":"6 Ltr",   "category":"Buttermilk", "emoji":"🥤"},
    {"id":"banas_buttermilk_500ml", "name":"Banas ",      "variant":"500 mL",  "category":"Buttermilk", "emoji":"🥤"},
    # Curd (full category)
    {"id":"dahi_cup_85gm",          "name":"Amul Dahi Cup",         "variant":"85 gm",   "category":"Curd",       "emoji":"🍶"},
    {"id":"dahi_cup_200gm",         "name":"Amul Dahi Cup",         "variant":"200 gm",  "category":"Curd",       "emoji":"🍶"},
    {"id":"dahi_1kg",               "name":"Amul Dahi",             "variant":"1 Kg",    "category":"Curd",       "emoji":"🍶"},
    {"id":"dahi_5kg",               "name":"Amul Dahi",             "variant":"5 Kg",    "category":"Curd",       "emoji":"🍶"},
    {"id":"lite_dahi_pouch",        "name":"Lite Dahi",             "variant":"Pouch",   "category":"Curd",       "emoji":"🍶"},
]

# ── Product column order for exports (fixed left-to-right order) ──────────────
EXPORT_PRODUCT_COLS = [
    # Milk
    ("amul_gold_500ml",        "Gold 500mL"),
    ("amul_gold_6ltr",         "Gold 6Ltr"),
    ("amul_taaza_500ml",       "Taaza 500mL"),
    ("amul_taaza_6ltr",        "Taaza 6Ltr"),
    ("amul_taaza_pouch",       "Taaza Pouch"),
    # Buttermilk
    ("buttermilk_500ml",       "BM 500mL"),
    ("buttermilk_6ltr",        "BM 6Ltr"),
    ("banas_buttermilk_500ml", "Banas BM 500mL"),
    # Curd
    ("dahi_cup_85gm",          "Dahi Cup 85gm"),
    ("dahi_cup_200gm",         "Dahi Cup 200gm"),
    ("dahi_1kg",               "Dahi 1Kg"),
    ("dahi_5kg",               "Dahi 5Kg"),
    ("lite_dahi_pouch",        "Lite Dahi Pouch"),
]

# ── Helpers ───────────────────────────────────────────────────────────────────
def hash_pin(pin): return hashlib.sha256(str(pin).encode()).hexdigest()

def load_json(path, default):
    try:
        if os.path.exists(path):
            with open(path, encoding="utf-8") as f: return json.load(f)
    except: pass
    return default

def save_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

def load_customers():  return load_json(CUSTOMERS_FILE, [])
def save_customers(d): save_json(CUSTOMERS_FILE, d)
def load_orders():     return load_json(ORDERS_FILE, [])
def save_orders(d):    save_json(ORDERS_FILE, d)
def load_notes():      return load_json(NOTES_FILE, {})
def save_notes(d):     save_json(NOTES_FILE, d)
def load_routes():     return load_json(ROUTES_FILE, [])
def save_routes(d):    save_json(ROUTES_FILE, d)
def load_push_subs():  return load_json(PUSH_SUBS_FILE, {})
def save_push_subs(d): save_json(PUSH_SUBS_FILE, d)

def load_admin():
    default = {"username":"admin","pin":hash_pin("1234"),"name":"Admin","whatsapp":""}
    if not os.path.exists(ADMIN_FILE): save_json(ADMIN_FILE, default)
    return load_json(ADMIN_FILE, default)

def strip_pin(c):
    return {k:v for k,v in c.items() if k != "pin"}

# ── CORS + session ────────────────────────────────────────────────────────────
@app.after_request
def add_cors(r):
    r.headers["Access-Control-Allow-Origin"]      = request.headers.get("Origin","*")
    r.headers["Access-Control-Allow-Credentials"] = "true"
    r.headers["Access-Control-Allow-Headers"]     = "Content-Type"
    r.headers["Access-Control-Allow-Methods"]     = "GET,POST,PUT,DELETE,OPTIONS"
    return r

@app.route("/api/<path:p>", methods=["OPTIONS"])
def opt(p): return "", 204

# ── Auth guards ───────────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def w(*a,**k):
        if not session.get("user_id"): return jsonify({"error":"Not authenticated"}), 401
        return f(*a,**k)
    return w

def admin_required(f):
    @wraps(f)
    def w(*a,**k):
        if not session.get("is_admin"): return jsonify({"error":"Admin only"}), 403
        return f(*a,**k)
    return w

# ── WhatsApp link builder ─────────────────────────────────────────────────────
def wa_link(phone, msg):
    clean = re.sub(r'\D','', str(phone))
    if not clean.startswith('91'): clean = '91' + clean
    return f"https://wa.me/{clean}?text={quote(msg)}"

def order_receipt_msg(order, lang="en"):
    lines = []
    date_str = order.get("date","")
    oid      = order.get("id","")
    if lang == "hi":
        lines.append(f"✅ *ऑर्डर कन्फर्म हो गया!*")
        lines.append(f"📋 ऑर्डर ID: {oid}")
        lines.append(f"📅 तारीख: {date_str}  🕐 समय: {order.get('time','')}")
        lines.append("")
        lines.append("*आपने जो मँगाया:*")
        for it in order.get("items",[]):
            lines.append(f"  • {it['product_name']} {it['variant']} × {it['quantity']}")
        if order.get("notes"): lines.append(f"\n📝 नोट: {order['notes']}")
        lines.append(f"\nधन्यवाद! 🙏 अगर कोई बदलाव चाहिए तो 2:30 बजे से पहले करें।")
    else:
        lines.append(f"✅ *Order Confirmed!*")
        lines.append(f"📋 Order ID: {oid}")
        lines.append(f"📅 Date: {date_str}  🕐 Time: {order.get('time','')}")
        lines.append("")
        lines.append("*Items ordered:*")
        for it in order.get("items",[]):
            lines.append(f"  • {it['product_name']} {it['variant']} × {it['quantity']}")
        if order.get("notes"): lines.append(f"\n📝 Note: {order['notes']}")
        lines.append(f"\nThank you! 🙏 You can edit before {cutoff_str()} IST.")
    return "\n".join(lines)

def missing_reminder_msg(name, lang="en"):
    if lang == "hi":
        return f"🥛 नमस्ते {name}!\n\nआज का ऑर्डर अभी तक नहीं दिया गया है। कृपया 2:30 बजे से पहले ऑर्डर करें।\n\nधन्यवाद! 🙏"
    return f"🥛 Hello {name}!\n\nYou haven't placed your order for today yet. Please order before {cutoff_str()} IST.\n\nThank you! 🙏"

# ── Pivot builder (shared between Excel & PDF) ────────────────────────────────
def build_pivot(orders):
    """
    Returns (pivot_rows, col_ids, col_labels) where:
      pivot_rows = [ {customer_name, area, route, phone, col_id: qty, ...}, ... ]
      col_ids    = list of product ids in export order (only those with any qty)
      col_labels = matching human-readable short labels
    """
    # Aggregate per customer
    cust_map = {}  # customer_id -> { meta, pid: qty }
    for o in orders:
        cid = o["customer_id"]
        if cid not in cust_map:
            cust_map[cid] = {
                "customer_name":  o["customer_name"],
                "area":           o.get("customer_area", ""),
                "route":          o.get("customer_route", ""),
                "phone":          o.get("customer_phone", ""),
            }
        for item in o.get("items", []):
            pid = item["product_id"]
            cust_map[cid][pid] = cust_map[cid].get(pid, 0) + item["quantity"]

    # Determine which columns have any data
    used_pids = set()
    for row in cust_map.values():
        for pid, _ in EXPORT_PRODUCT_COLS:
            if row.get(pid, 0) > 0:
                used_pids.add(pid)

    col_ids    = [pid for pid, _ in EXPORT_PRODUCT_COLS if pid in used_pids]
    col_labels = [lbl for pid, lbl in EXPORT_PRODUCT_COLS if pid in used_pids]

    # Sort rows by route then customer name
    pivot_rows = sorted(cust_map.values(), key=lambda r: (r.get("route",""), r["customer_name"]))
    return pivot_rows, col_ids, col_labels


# ── Serve HTML ────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    with open("templates/index.html", encoding="utf-8") as f: return f.read()

@app.route("/sw.js")
def sw():
    with open("templates/sw.js", encoding="utf-8") as f:
        return Response(f.read(), mimetype="application/javascript")

# ══════════════════════════════════════════════════════════════════════════════
#  AUTH
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/api/auth/login", methods=["POST"])
def login():
    data  = request.json or {}
    phone = data.get("phone","").strip()
    pin   = str(data.get("pin","")).strip()
    remember = data.get("remember", False)

    admin = load_admin()
    if phone == admin["username"] and hash_pin(pin) == admin["pin"]:
        session.permanent = True
        session["user_id"]   = "admin"
        session["user_name"] = admin["name"]
        session["is_admin"]  = True
        return jsonify({"success":True,"is_admin":True,"name":admin["name"]})

    for c in load_customers():
        if c["phone"] == phone and hash_pin(pin) == c["pin"]:
            session.permanent = True
            session["user_id"]   = c["id"]
            session["user_name"] = c["name"]
            session["is_admin"]  = False
            resp = jsonify({"success":True,"is_admin":False,"name":c["name"],
                            "customer":strip_pin(c)})
            if remember:
                resp.set_cookie("remember_id", c["id"], max_age=30*24*3600, httponly=True, samesite="Lax")
                resp.set_cookie("remember_pin", hash_pin(pin), max_age=30*24*3600, httponly=True, samesite="Lax")
            return resp

    return jsonify({"success":False,"error":"Wrong phone number or PIN. Try again."}), 401

@app.route("/api/auth/quick-login", methods=["POST"])
def quick_login():
    rid  = request.cookies.get("remember_id","")
    rpin = request.cookies.get("remember_pin","")
    if not rid or not rpin:
        return jsonify({"success":False,"error":"No saved login"}), 401
    for c in load_customers():
        if c["id"] == rid and hash_pin("") != rpin and c["pin"] == rpin:
            session.permanent = True
            session["user_id"]   = c["id"]
            session["user_name"] = c["name"]
            session["is_admin"]  = False
            return jsonify({"success":True,"is_admin":False,"name":c["name"],"customer":strip_pin(c)})
    return jsonify({"success":False,"error":"Session expired"}), 401

@app.route("/api/auth/logout", methods=["POST"])
def logout():
    session.clear()
    resp = jsonify({"success":True})
    resp.delete_cookie("remember_id")
    resp.delete_cookie("remember_pin")
    return resp

@app.route("/api/auth/me", methods=["GET"])
def me():
    if not session.get("user_id"): return jsonify({"logged_in":False})
    c = next((x for x in load_customers() if x["id"]==session["user_id"]), None)
    return jsonify({"logged_in":True,"user_id":session["user_id"],
                    "name":session["user_name"],"is_admin":session.get("is_admin",False),
                    "customer":strip_pin(c) if c else None})

@app.route("/api/auth/change-pin", methods=["POST"])
@login_required
def change_pin():
    data    = request.json or {}
    old_pin = str(data.get("old_pin",""))
    new_pin = str(data.get("new_pin",""))
    if len(new_pin)!=4 or not new_pin.isdigit():
        return jsonify({"success":False,"error":"PIN must be exactly 4 digits"}), 400
    if session.get("is_admin"):
        admin = load_admin()
        if hash_pin(old_pin) != admin["pin"]: return jsonify({"success":False,"error":"Old PIN incorrect"}), 400
        admin["pin"] = hash_pin(new_pin)
        save_json(ADMIN_FILE, admin)
        return jsonify({"success":True})
    customers = load_customers()
    for c in customers:
        if c["id"] == session["user_id"]:
            if hash_pin(old_pin) != c["pin"]: return jsonify({"success":False,"error":"Old PIN incorrect"}), 400
            c["pin"] = hash_pin(new_pin)
            save_customers(customers)
            return jsonify({"success":True})
    return jsonify({"success":False,"error":"User not found"}), 404

# ══════════════════════════════════════════════════════════════════════════════
#  ORDER STATUS
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/api/order-status", methods=["GET"])
@login_required
def order_status():
    n      = now_ist()
    open_  = is_order_open()
    cutoff = n.replace(hour=ORDER_CUTOFF_HOUR, minute=ORDER_CUTOFF_MINUTE, second=0, microsecond=0)
    mins   = max(0,int((cutoff-n).total_seconds()//60)) if open_ else 0
    return jsonify({"open":open_,"cutoff_time":cutoff_str(),"current_ist":n.strftime("%H:%M"),
                    "minutes_left":mins,"is_admin":session.get("is_admin",False)})

# ══════════════════════════════════════════════════════════════════════════════
#  PRODUCTS
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/api/products", methods=["GET"])
@login_required
def get_products(): return jsonify(PRODUCTS)

# ══════════════════════════════════════════════════════════════════════════════
#  ROUTES
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/api/routes", methods=["GET"])
@login_required
def get_routes(): return jsonify(load_routes())

@app.route("/api/routes", methods=["POST"])
@login_required
@admin_required
def save_route():
    routes = load_routes()
    data   = request.json or {}
    name   = data.get("name","").strip()
    if not name: return jsonify({"success":False,"error":"Route name required"}), 400
    if name not in routes: routes.append(name)
    save_routes(routes)
    return jsonify({"success":True,"routes":routes})

@app.route("/api/routes/<name>", methods=["DELETE"])
@login_required
@admin_required
def delete_route(name):
    routes = [r for r in load_routes() if r != name]
    save_routes(routes)
    return jsonify({"success":True})

# ══════════════════════════════════════════════════════════════════════════════
#  CUSTOMERS
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/api/customers", methods=["GET"])
@login_required
def get_customers():
    return jsonify([strip_pin(c) for c in load_customers()])

@app.route("/api/customers", methods=["POST"])
@login_required
@admin_required
def add_customer():
    data  = request.json or {}
    name  = data.get("name","").strip()
    phone = data.get("phone","").strip()
    pin   = str(data.get("pin","")).strip()
    area  = data.get("area","").strip()
    route = data.get("route","").strip()
    lang  = data.get("lang","en")
    if not name or not phone or not pin:
        return jsonify({"success":False,"error":"Name, phone and PIN required"}), 400
    if len(pin)!=4 or not pin.isdigit():
        return jsonify({"success":False,"error":"PIN must be 4 digits"}), 400
    customers = load_customers()
    if any(c["phone"]==phone for c in customers):
        return jsonify({"success":False,"error":"Phone already registered"}), 400
    cid = f"C{len(customers)+1:03d}_{int(now_ist().timestamp())}"
    c   = {"id":cid,"name":name,"phone":phone,"pin":hash_pin(pin),"area":area,
           "route":route,"lang":lang,"created_at":now_ist().isoformat(),"streak":0,"last_order_date":""}
    customers.append(c)
    save_customers(customers)
    return jsonify({"success":True,"customer":strip_pin(c)})

@app.route("/api/customers/<cid>", methods=["PUT"])
@login_required
@admin_required
def update_customer(cid):
    data = request.json or {}
    customers = load_customers()
    for c in customers:
        if c["id"] == cid:
            for f in ["name","phone","area","route","lang"]:
                if f in data: c[f] = data[f]
            if data.get("pin"):
                pin = str(data["pin"])
                if len(pin)!=4 or not pin.isdigit():
                    return jsonify({"success":False,"error":"PIN must be 4 digits"}), 400
                c["pin"] = hash_pin(pin)
            save_customers(customers)
            return jsonify({"success":True})
    return jsonify({"success":False,"error":"Not found"}), 404

@app.route("/api/customers/<cid>", methods=["DELETE"])
@login_required
@admin_required
def delete_customer(cid):
    save_customers([c for c in load_customers() if c["id"]!=cid])
    return jsonify({"success":True})

@app.route("/api/admin/reset-customer-pin", methods=["POST"])
@login_required
@admin_required
def reset_customer_pin():
    data    = request.json or {}
    cid     = data.get("customer_id")
    new_pin = str(data.get("new_pin",""))
    if len(new_pin)!=4 or not new_pin.isdigit():
        return jsonify({"success":False,"error":"PIN must be 4 digits"}), 400
    customers = load_customers()
    for c in customers:
        if c["id"]==cid:
            c["pin"] = hash_pin(new_pin)
            save_customers(customers)
            return jsonify({"success":True})
    return jsonify({"success":False,"error":"Not found"}), 404

# ══════════════════════════════════════════════════════════════════════════════
#  CUSTOMER NOTES
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/api/notes/<cid>", methods=["GET"])
@login_required
@admin_required
def get_note(cid):
    notes = load_notes()
    return jsonify({"note":notes.get(cid,"")})

@app.route("/api/notes/<cid>", methods=["POST"])
@login_required
@admin_required
def save_note(cid):
    notes = load_notes()
    notes[cid] = (request.json or {}).get("note","")
    save_notes(notes)
    return jsonify({"success":True})

# ══════════════════════════════════════════════════════════════════════════════
#  ORDERS
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/api/orders", methods=["POST"])
@login_required
def place_order():
    if not session.get("is_admin") and not is_order_open():
        return jsonify({"success":False,"error":f"Order window closed. Please order before {cutoff_str()} IST."}), 403

    data      = request.json or {}
    orders    = load_orders()
    customers = load_customers()

    if session.get("is_admin"):
        cid = data.get("customer_id")
        c   = next((x for x in customers if x["id"]==cid), None)
        if not c: return jsonify({"success":False,"error":"Customer not found"}), 400
    else:
        cid = session["user_id"]
        c   = next((x for x in customers if x["id"]==cid), None)

    n   = now_ist()
    oid = f"ORD{n.strftime('%Y%m%d%H%M%S')}{len(orders)+1:03d}"
    order = {
        "id":oid,"customer_id":c["id"] if c else "admin",
        "customer_name":c["name"] if c else "Admin",
        "customer_phone":c.get("phone","") if c else "",
        "customer_area":c.get("area","") if c else "",
        "customer_route":c.get("route","") if c else "",
        "items":data["items"],"notes":data.get("notes",""),
        "status":"pending","created_at":n.isoformat(),
        "date":n.strftime("%d-%m-%Y"),"time":n.strftime("%H:%M"),
    }
    orders.append(order)
    save_orders(orders)

    if c:
        today_str = n.strftime("%d-%m-%Y")
        for cu in customers:
            if cu["id"] == c["id"]:
                if cu.get("last_order_date") != today_str:
                    cu["streak"] = cu.get("streak",0) + 1
                cu["last_order_date"] = today_str
                break
        save_customers(customers)

    receipt = order_receipt_msg(order, c.get("lang","en") if c else "en")
    wa = wa_link(c["phone"], receipt) if c else ""
    return jsonify({"success":True,"order_id":oid,"order":order,"wa_receipt_link":wa,"receipt_msg":receipt})

@app.route("/api/orders", methods=["GET"])
@login_required
def get_orders():
    orders = load_orders()
    date_filter  = request.args.get("date","")
    route_filter = request.args.get("route","")
    if not session.get("is_admin"):
        orders = [o for o in orders if o["customer_id"]==session["user_id"]]
    if date_filter:  orders = [o for o in orders if o.get("date")==date_filter]
    if route_filter: orders = [o for o in orders if o.get("customer_route")==route_filter]
    return jsonify(sorted(orders, key=lambda x:x.get("created_at",""), reverse=True))

@app.route("/api/orders/<oid>", methods=["PUT"])
@login_required
def update_order(oid):
    data   = request.json or {}
    orders = load_orders()
    for o in orders:
        if o["id"] == oid:
            if not session.get("is_admin"):
                if o["customer_id"] != session["user_id"]: return jsonify({"error":"Forbidden"}), 403
                if o["status"] != "pending": return jsonify({"success":False,"error":"Cannot edit confirmed order"}), 400
                if not is_order_open(): return jsonify({"success":False,"error":f"Edit window closed ({cutoff_str()} IST)"}), 403
            o["items"]      = data.get("items", o["items"])
            o["notes"]      = data.get("notes", o.get("notes",""))
            o["status"]     = data.get("status", o["status"])
            o["updated_at"] = now_ist().isoformat()
            save_orders(orders)
            return jsonify({"success":True,"order":o})
    return jsonify({"success":False,"error":"Not found"}), 404

@app.route("/api/orders/<oid>", methods=["DELETE"])
@login_required
@admin_required
def delete_order(oid):
    save_orders([o for o in load_orders() if o["id"]!=oid])
    return jsonify({"success":True})

@app.route("/api/orders/duplicate-yesterday", methods=["POST"])
@login_required
def duplicate_yesterday():
    if not is_order_open() and not session.get("is_admin"):
        return jsonify({"success":False,"error":f"Order window closed"}), 403
    n         = now_ist()
    yesterday = (n - timedelta(days=1)).strftime("%d-%m-%Y")
    today     = n.strftime("%d-%m-%Y")
    orders    = load_orders()
    cid       = session["user_id"]

    yesterday_orders = [o for o in orders if o["customer_id"]==cid and o.get("date")==yesterday and o.get("status")!="cancelled"]
    if not yesterday_orders: return jsonify({"success":False,"error":"No order found for yesterday"}), 404

    already_today = any(o for o in orders if o["customer_id"]==cid and o.get("date")==today and o.get("status")!="cancelled")
    if already_today: return jsonify({"success":False,"error":"You already placed an order today"}), 400

    src  = yesterday_orders[0]
    oid  = f"ORD{n.strftime('%Y%m%d%H%M%S')}{len(orders)+1:03d}"
    new_order = {**src, "id":oid,"date":today,"time":n.strftime("%H:%M"),
                 "created_at":n.isoformat(),"status":"pending","notes":src.get("notes","")}
    orders.append(new_order)
    save_orders(orders)
    return jsonify({"success":True,"order_id":oid,"order":new_order})

@app.route("/api/orders/<oid>/wa-confirm", methods=["GET"])
@login_required
@admin_required
def wa_confirm_link(oid):
    orders = load_orders()
    o = next((x for x in orders if x["id"]==oid), None)
    if not o: return jsonify({"error":"Not found"}), 404
    customers = load_customers()
    c = next((x for x in customers if x["id"]==o["customer_id"]), None)
    lang = c.get("lang","en") if c else "en"
    msg  = order_receipt_msg(o, lang)
    link = wa_link(o["customer_phone"], msg) if o["customer_phone"] else ""
    return jsonify({"link":link,"message":msg})

@app.route("/api/missing-orders/wa-remind/<cid>", methods=["GET"])
@login_required
@admin_required
def wa_remind(cid):
    customers = load_customers()
    c = next((x for x in customers if x["id"]==cid), None)
    if not c: return jsonify({"error":"Not found"}), 404
    msg  = missing_reminder_msg(c["name"], c.get("lang","en"))
    link = wa_link(c["phone"], msg)
    return jsonify({"link":link})

# ══════════════════════════════════════════════════════════════════════════════
#  SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/api/summary", methods=["GET"])
@login_required
@admin_required
def get_summary():
    orders = load_orders()
    date_filter  = request.args.get("date","")
    route_filter = request.args.get("route","")
    if date_filter:  orders = [o for o in orders if o.get("date")==date_filter]
    if route_filter: orders = [o for o in orders if o.get("customer_route")==route_filter]
    orders = [o for o in orders if o.get("status")!="cancelled"]
    summary = {}
    for o in orders:
        for item in o.get("items",[]):
            pid = item["product_id"]
            if pid not in summary:
                summary[pid]={**item,"total_quantity":0,"customer_count":0,"customers":set()}
            summary[pid]["total_quantity"] += item["quantity"]
            summary[pid]["customers"].add(o["customer_id"])
    return jsonify({"summary":[{**s,"customer_count":len(s["customers"]),"customers":list(s["customers"])} for s in summary.values()],
                    "total_orders":len(orders),"date":date_filter})

# ══════════════════════════════════════════════════════════════════════════════
#  MISSING ORDERS
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/api/missing-orders", methods=["GET"])
@login_required
@admin_required
def missing_orders():
    date_filter  = request.args.get("date", now_ist().strftime("%d-%m-%Y"))
    route_filter = request.args.get("route","")
    orders    = load_orders()
    customers = load_customers()
    ordered_ids = set(o["customer_id"] for o in orders if o.get("date")==date_filter and o.get("status")!="cancelled")
    missing = [{"id":c["id"],"name":c["name"],"phone":c["phone"],"area":c.get("area",""),"route":c.get("route","")}
               for c in customers if c["id"] not in ordered_ids
               and (not route_filter or c.get("route","")==route_filter)]
    return jsonify({"date":date_filter,"total_customers":len(customers),"ordered_count":len(ordered_ids),
                    "missing_count":len(missing),"missing":missing})

# ══════════════════════════════════════════════════════════════════════════════
#  ANALYTICS
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/api/analytics", methods=["GET"])
@login_required
@admin_required
def analytics():
    orders    = load_orders()
    customers = load_customers()
    n         = now_ist()

    daily = {}
    for i in range(13, -1, -1):
        d = (n - timedelta(days=i)).strftime("%d-%m-%Y")
        daily[d] = 0
    for o in orders:
        if o.get("date") in daily and o.get("status") != "cancelled":
            daily[o["date"]] += 1

    cutoff30 = n - timedelta(days=30)
    prod_totals = {}
    for o in orders:
        try:
            od = datetime.strptime(o["date"],"%d-%m-%Y").replace(tzinfo=IST)
        except: continue
        if od < cutoff30 or o.get("status")=="cancelled": continue
        for item in o.get("items",[]):
            k = f"{item['product_name']} {item['variant']}"
            prod_totals[k] = prod_totals.get(k,0) + item["quantity"]

    top_streaks = sorted(customers, key=lambda c: c.get("streak",0), reverse=True)[:5]

    total_custs = len(customers)
    completion = {}
    for i in range(6, -1, -1):
        d = (n - timedelta(days=i)).strftime("%d-%m-%Y")
        ordered = len(set(o["customer_id"] for o in orders if o.get("date")==d and o.get("status")!="cancelled"))
        completion[d] = round((ordered/total_custs*100) if total_custs else 0, 1)

    return jsonify({
        "daily_orders": [{"date":k,"count":v} for k,v in daily.items()],
        "product_totals": [{"product":k,"qty":v} for k,v in sorted(prod_totals.items(),key=lambda x:-x[1])[:8]],
        "top_streaks": [{"name":c["name"],"streak":c.get("streak",0)} for c in top_streaks],
        "completion_rate": [{"date":k,"rate":v} for k,v in completion.items()],
    })

# ══════════════════════════════════════════════════════════════════════════════
#  PUSH NOTIFICATIONS
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/api/push/subscribe", methods=["POST"])
@login_required
def push_subscribe():
    data = request.json or {}
    subs = load_push_subs()
    uid  = session["user_id"]
    subs[uid] = data.get("subscription",{})
    save_push_subs(subs)
    return jsonify({"success":True})

@app.route("/api/push/test", methods=["POST"])
@login_required
@admin_required
def push_test():
    return jsonify({"success":True,"note":"Push notification triggered. Install pywebpush on server for real delivery."})

# ══════════════════════════════════════════════════════════════════════════════
#  DRIVER SHEET
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/api/driver-sheet", methods=["GET"])
@login_required
@admin_required
def driver_sheet():
    date_filter  = request.args.get("date", now_ist().strftime("%d-%m-%Y"))
    route_filter = request.args.get("route","")
    orders  = [o for o in load_orders() if o.get("date")==date_filter and o.get("status")!="cancelled"]
    if route_filter: orders = [o for o in orders if o.get("customer_route")==route_filter]
    orders  = sorted(orders, key=lambda o:(o.get("customer_route","zzz"),o["customer_name"]))

    rows_html = ""
    for idx, o in enumerate(orders,1):
        items_str = ", ".join(f"{i['product_name']} {i['variant']}×{i['quantity']}" for i in o["items"])
        rows_html += f"""<tr>
          <td>{idx}</td>
          <td><strong>{o['customer_name']}</strong><br><small>{o.get('customer_area','')}</small></td>
          <td style="color:#003087;font-weight:700">{o.get('customer_route','—')}</td>
          <td>{items_str}</td>
          <td>{o.get('notes','')}</td>
          <td style="width:80px"></td>
        </tr>"""

    html = f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
    <title>Driver Sheet — {date_filter}</title>
    <style>
      body{{font-family:Arial,sans-serif;font-size:13px;margin:20px}}
      h1{{color:#003087;margin-bottom:4px}}
      .meta{{color:#666;margin-bottom:16px;font-size:12px}}
      table{{width:100%;border-collapse:collapse}}
      th{{background:#003087;color:#fff;padding:8px 10px;text-align:left;font-size:12px}}
      td{{padding:8px 10px;border-bottom:1px solid #E2E8F0;vertical-align:top}}
      tr:nth-child(even){{background:#F8FAFF}}
      .footer{{margin-top:20px;font-size:11px;color:#999;text-align:center}}
      @media print{{body{{margin:0}}}}
    </style></head><body>
    <h1>🚛 Delivery Sheet — {date_filter}</h1>
    <div class="meta">Route: {route_filter or 'All Routes'} &nbsp;|&nbsp; Total: {len(orders)} customers &nbsp;|&nbsp; Generated: {now_ist().strftime('%H:%M IST')}</div>
    <table>
      <thead><tr><th>#</th><th>Customer</th><th>Route</th><th>Items to Deliver</th><th>Notes</th><th>✓ Delivered</th></tr></thead>
      <tbody>{rows_html}</tbody>
    </table>
    <div class="footer">Amul Wholesale — {date_filter}</div>
    <script>window.onload=()=>window.print()</script>
    </body></html>"""
    return Response(html, mimetype="text/html")

# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT  — PIVOT FORMAT (one row per customer, products as columns)
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/api/export/excel", methods=["GET"])
@login_required
@admin_required
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    date_filter  = request.args.get("date", now_ist().strftime("%d-%m-%Y"))
    route_filter = request.args.get("route", "")

    all_orders = load_orders()
    orders = [o for o in all_orders if o.get("date") == date_filter and o.get("status") != "cancelled"]
    if route_filter:
        orders = [o for o in orders if o.get("customer_route") == route_filter]

    customers = load_customers()
    ordered_ids = {o["customer_id"] for o in orders}
    missing = [
        {"name": c["name"], "phone": c["phone"], "area": c.get("area", ""), "route": c.get("route", "")}
        for c in customers
        if c["id"] not in ordered_ids and (not route_filter or c.get("route", "") == route_filter)
    ]

    pivot_rows, col_ids, col_labels = build_pivot(orders)

    # Total Summary (product-wise)
    prod_meta = {p["id"]: p for p in PRODUCTS}
    summary_map = {}  # pid -> {product_name, variant, category, total_qty, customers:set()}
    for o in orders:
        for item in o.get("items", []):
            pid = item["product_id"]
            if pid not in summary_map:
                pm = prod_meta.get(pid, {})
                summary_map[pid] = {
                    "product_name": item.get("product_name", pm.get("name", "")),
                    "variant": item.get("variant", pm.get("variant", "")),
                    "category": pm.get("category", ""),
                    "total_qty": 0,
                    "customers": set(),
                }
            summary_map[pid]["total_qty"] += int(item.get("quantity", 0) or 0)
            summary_map[pid]["customers"].add(o["customer_id"])
    summary_rows = sorted(
        [
            {
                "product_name": v["product_name"],
                "variant": v["variant"],
                "category": v["category"],
                "total_qty": v["total_qty"],
                "customer_count": len(v["customers"]),
            }
            for v in summary_map.values()
            if v["total_qty"] > 0
        ],
        key=lambda r: (r.get("category", ""), r.get("product_name", ""), r.get("variant", "")),
    )

    # Route-wise Breakdown (pivot: one row per customer, sorted by route)
    breakdown_rows = pivot_rows

    wb = Workbook()

    # ── Styles: clean/professional ─────────────────────────────────
    header_fill = PatternFill("solid", fgColor="003087")
    title_fill  = PatternFill("solid", fgColor="F1F5FB")
    zebra_fill  = PatternFill("solid", fgColor="F8FAFC")
    white_fill  = PatternFill("solid", fgColor="FFFFFF")

    hfont = Font(bold=True, color="FFFFFF", size=10)
    title_font = Font(bold=True, size=14, color="003087")
    sub_font = Font(italic=True, color="666666", size=9)
    name_font = Font(bold=True, size=10, color="0D1B2A")
    body_font = Font(size=9, color="0D1B2A")

    thin = Side(style="thin", color="D0D7E2")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ══════════════════════════════════════════════════════════════
    # Sheet 1 — Order Sheet (pivot)
    # ══════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Order Sheet"

    total_cols = 4 + len(col_ids) + 1
    last_col_letter = get_column_letter(total_cols)
    ws.freeze_panes = "E4"

    title_txt = f"Amul Daily Order Sheet — {date_filter}"
    if route_filter:
        title_txt += f"  |  Route: {route_filter}"
    ws.merge_cells(f"A1:{last_col_letter}1")
    c = ws["A1"]
    c.value = title_txt
    c.font = title_font
    c.fill = title_fill
    c.alignment = center
    ws.row_dimensions[1].height = 24

    ws.merge_cells(f"A2:{last_col_letter}2")
    c = ws["A2"]
    c.value = f"Generated: {now_ist().strftime('%d-%m-%Y %H:%M')} IST   |   Customers ordered: {len(pivot_rows)}   |   Not ordered: {len(missing)}"
    c.font = sub_font
    c.alignment = center

    ws.row_dimensions[3].height = 30
    header_row = ["#", "Customer Name", "Route", "Phone"] + col_labels + ["Total (₹)"]
    for col_idx, hdr in enumerate(header_row, 1):
        cell = ws.cell(row=3, column=col_idx, value=hdr)
        cell.font = hfont
        cell.fill = header_fill
        cell.alignment = center
        cell.border = bdr

    total_col_idx = len(header_row)
    for i, row_data in enumerate(pivot_rows, 1):
        r = 3 + i
        ws.row_dimensions[r].height = 20
        fill = zebra_fill if i % 2 == 0 else white_fill

        meta_vals = [i, row_data["customer_name"], row_data.get("route", "—"), row_data.get("phone", "")]
        for col_idx, val in enumerate(meta_vals, 1):
            c = ws.cell(row=r, column=col_idx, value=val)
            c.fill = fill
            c.border = bdr
            c.alignment = left if col_idx == 2 else center
            c.font = name_font if col_idx == 2 else body_font

        for j, pid in enumerate(col_ids):
            qty = int(row_data.get(pid, 0) or 0)
            col = 5 + j
            c = ws.cell(row=r, column=col, value=(qty if qty > 0 else ""))
            c.fill = fill
            c.border = bdr
            c.alignment = center
            c.font = Font(bold=True, size=10, color="003087") if qty > 0 else body_font

        c = ws.cell(row=r, column=total_col_idx, value="")
        c.fill = fill
        c.border = bdr
        c.alignment = center

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    for i in range(len(col_ids)):
        ws.column_dimensions[get_column_letter(5 + i)].width = 12
    ws.column_dimensions[get_column_letter(total_col_idx)].width = 12

    # ══════════════════════════════════════════════════════════════
    # Sheet 2 — Total Summary
    # ══════════════════════════════════════════════════════════════
    ws_sum = wb.create_sheet("Total Summary")
    ws_sum.freeze_panes = "A4"
    ws_sum.merge_cells("A1:E1")
    c = ws_sum["A1"]
    c.value = f"Total Summary — {date_filter}" + (f"  |  Route: {route_filter}" if route_filter else "")
    c.font = title_font
    c.fill = title_fill
    c.alignment = center
    ws_sum.merge_cells("A2:E2")
    c = ws_sum["A2"]
    c.value = f"Generated: {now_ist().strftime('%d-%m-%Y %H:%M')} IST"
    c.font = sub_font
    c.alignment = center

    sum_headers = ["Product", "Variant", "Category", "Total Qty", "# Customers"]
    for col_idx, hdr in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=3, column=col_idx, value=hdr)
        cell.font = hfont
        cell.fill = header_fill
        cell.alignment = center
        cell.border = bdr

    for i, s in enumerate(summary_rows, 1):
        r = 3 + i
        fill = zebra_fill if i % 2 == 0 else white_fill
        vals = [s["product_name"], s["variant"], s["category"], s["total_qty"], s["customer_count"]]
        for col_idx, val in enumerate(vals, 1):
            c = ws_sum.cell(row=r, column=col_idx, value=val)
            c.fill = fill
            c.border = bdr
            c.alignment = left if col_idx in (1, 2, 3) else center
            c.font = body_font

    for i, w in enumerate([26, 14, 12, 10, 12], 1):
        ws_sum.column_dimensions[get_column_letter(i)].width = w

    # ══════════════════════════════════════════════════════════════
    # Sheet 3 — Route-wise Breakdown
    # ══════════════════════════════════════════════════════════════
    ws_rb = wb.create_sheet("Route-wise Breakdown")
    ws_rb.freeze_panes = "A4"
    # Columns: Route, Customer, Area, Phone + product columns
    rb_total_cols = 4 + len(col_ids)
    rb_last_col_letter = get_column_letter(rb_total_cols)
    ws_rb.merge_cells(f"A1:{rb_last_col_letter}1")
    c = ws_rb["A1"]
    c.value = f"Route-wise Breakdown — {date_filter}" + (f"  |  Route: {route_filter}" if route_filter else "")
    c.font = title_font
    c.fill = title_fill
    c.alignment = center
    ws_rb.merge_cells(f"A2:{rb_last_col_letter}2")
    c = ws_rb["A2"]
    c.value = f"Generated: {now_ist().strftime('%d-%m-%Y %H:%M')} IST"
    c.font = sub_font
    c.alignment = center

    rb_headers = ["Route", "Customer", "Area", "Phone"] + col_labels
    for col_idx, hdr in enumerate(rb_headers, 1):
        cell = ws_rb.cell(row=3, column=col_idx, value=hdr)
        cell.font = hfont
        cell.fill = header_fill
        cell.alignment = center
        cell.border = bdr

    for i, rrow in enumerate(breakdown_rows, 1):
        r = 3 + i
        fill = zebra_fill if i % 2 == 0 else white_fill
        # Meta columns
        meta_vals = [rrow.get("route", "—"), rrow.get("customer_name", ""), rrow.get("area", ""), rrow.get("phone", "")]
        for col_idx, val in enumerate(meta_vals, 1):
            c = ws_rb.cell(row=r, column=col_idx, value=val)
            c.fill = fill
            c.border = bdr
            c.alignment = left if col_idx in (2, 3) else center
            c.font = name_font if col_idx == 2 else body_font

        # Product quantity columns
        for j, pid in enumerate(col_ids):
            qty = int(rrow.get(pid, 0) or 0)
            col = 5 + j
            c = ws_rb.cell(row=r, column=col, value=(qty if qty > 0 else ""))
            c.fill = fill
            c.border = bdr
            c.alignment = center
            c.font = Font(bold=True, size=10, color="003087") if qty > 0 else body_font

    ws_rb.column_dimensions["A"].width = 14
    ws_rb.column_dimensions["B"].width = 26
    ws_rb.column_dimensions["C"].width = 16
    ws_rb.column_dimensions["D"].width = 14
    for i in range(len(col_ids)):
        ws_rb.column_dimensions[get_column_letter(5 + i)].width = 12

    # ══════════════════════════════════════════════════════════════
    # Sheet 4 — Not Ordered Yet
    # ══════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Not Ordered Yet")
    ws2.freeze_panes = "A5"
    ws2.merge_cells("A1:F1")
    c = ws2["A1"]
    c.value = f"Customers Not Ordered Yet — {date_filter}" + (f"  |  Route: {route_filter}" if route_filter else "")
    c.font = title_font
    c.fill = title_fill
    c.alignment = center
    ws2.merge_cells("A2:F2")
    c = ws2["A2"]
    c.value = f"Generated: {now_ist().strftime('%d-%m-%Y %H:%M')} IST   |   Missing: {len(missing)}"
    c.font = sub_font
    c.alignment = center

    hrow2 = ["#", "Customer Name", "Phone", "Area", "Route", "Status"]
    for col_i, h in enumerate(hrow2, 1):
        c = ws2.cell(row=4, column=col_i, value=h)
        c.font = hfont
        c.fill = header_fill
        c.border = bdr
        c.alignment = center

    for ri, m in enumerate(missing, 1):
        r = 4 + ri
        fill = zebra_fill if ri % 2 == 0 else white_fill
        vals = [ri, m["name"], m["phone"], m["area"], m.get("route", ""), "Not Ordered"]
        for col_i, val in enumerate(vals, 1):
            c = ws2.cell(row=r, column=col_i, value=val)
            c.fill = fill
            c.border = bdr
            c.alignment = left if col_i in (2, 4, 5) else center
            c.font = Font(bold=True, color="E8001D", size=9) if col_i == 6 else body_font

    for i, w in enumerate([5, 26, 14, 16, 14, 14], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    fname = f"Amul_Orders_{date_filter.replace('-','_')}.xlsx"
    fpath = f"/tmp/{fname}"
    wb.save(fpath)
    return send_file(
        fpath,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ══════════════════════════════════════════════════════════════════════════════
#  PDF EXPORT  — PIVOT FORMAT (printable, same layout as Excel)
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/api/export/pdf", methods=["GET"])
@login_required
@admin_required
def export_pdf():
    date_filter  = request.args.get("date", now_ist().strftime("%d-%m-%Y"))
    route_filter = request.args.get("route", "")

    all_orders = load_orders()
    orders = [o for o in all_orders if o.get("date") == date_filter and o.get("status") != "cancelled"]
    if route_filter:
        orders = [o for o in orders if o.get("customer_route") == route_filter]

    pivot_rows, col_ids, col_labels = build_pivot(orders)

    customers = load_customers()
    ordered_ids = {o["customer_id"] for o in orders}
    missing = [
        {"name": c["name"], "phone": c["phone"], "area": c.get("area", ""), "route": c.get("route", "")}
        for c in customers
        if c["id"] not in ordered_ids and (not route_filter or c.get("route", "") == route_filter)
    ]

    # Total Summary (product-wise)
    prod_meta = {p["id"]: p for p in PRODUCTS}
    summary_map = {}
    for o in orders:
        for item in o.get("items", []):
            pid = item["product_id"]
            if pid not in summary_map:
                pm = prod_meta.get(pid, {})
                summary_map[pid] = {
                    "product_name": item.get("product_name", pm.get("name", "")),
                    "variant": item.get("variant", pm.get("variant", "")),
                    "category": pm.get("category", ""),
                    "total_qty": 0,
                    "customers": set(),
                }
            summary_map[pid]["total_qty"] += int(item.get("quantity", 0) or 0)
            summary_map[pid]["customers"].add(o["customer_id"])
    summary_rows = sorted(
        [
            {
                "product_name": v["product_name"],
                "variant": v["variant"],
                "category": v["category"],
                "total_qty": v["total_qty"],
                "customer_count": len(v["customers"]),
            }
            for v in summary_map.values()
            if v["total_qty"] > 0
        ],
        key=lambda r: (r.get("category", ""), r.get("product_name", ""), r.get("variant", "")),
    )

    # Route-wise Breakdown (pivot: one row per customer, sorted by route)
    breakdown_rows = pivot_rows

    # Main order sheet table (single header row; NO TOTAL QTY row)
    header_html = "<tr><th>#</th><th>Customer</th><th>Route</th><th>Phone</th>"
    for lbl in col_labels:
        header_html += f"<th>{lbl}</th>"
    header_html += "<th>Total (₹)</th></tr>"

    data_rows_html = ""
    for ri, row_data in enumerate(pivot_rows, 1):
        zebra = "zebra" if ri % 2 == 0 else ""
        cells = f"<td class='num'>{ri}</td>"
        cells += f"<td class='name'>{row_data['customer_name']}</td>"
        cells += f"<td class='num'>{row_data.get('route','—')}</td>"
        cells += f"<td class='num'>{row_data.get('phone','')}</td>"
        for pid in col_ids:
            qty = int(row_data.get(pid, 0) or 0)
            cells += f"<td class='qty'>{qty if qty > 0 else ''}</td>"
        cells += "<td class='num'></td>"
        data_rows_html += f"<tr class='{zebra}'>{cells}</tr>"

    # Summary table
    route_txt = route_filter or "All Routes"
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Amul Order Sheet — {date_filter}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: Arial, sans-serif; font-size: 11px; color: #0D1B2A; background: #fff; padding: 16px; }}
  .header {{ display: flex; align-items: center; justify-content: space-between; margin-bottom: 12px; border-bottom: 3px solid #003087; padding-bottom: 10px; }}
  .header-left h1 {{ font-size: 18px; font-weight: 900; color: #003087; }}
  .header-left p  {{ font-size: 11px; color: #64748B; margin-top: 3px; }}
  .header-right   {{ text-align: right; font-size: 11px; color: #64748B; }}
  .section-title  {{ font-size: 12px; font-weight: 800; color: #003087; margin: 14px 0 8px; text-transform: uppercase; letter-spacing: .4px; }}
  table  {{ width: 100%; border-collapse: collapse; font-size: 10px; }}
  th     {{ background: #003087; color: #fff; padding: 6px 6px; text-align: center; border: 1px solid #D0D7E2; font-size: 10px; }}
  td     {{ padding: 5px 6px; border: 1px solid #E2E8F0; vertical-align: middle; }}
  tr.zebra td {{ background: #F8FAFC; }}
  td.num {{ text-align: center; white-space: nowrap; }}
  td.name {{ font-weight: 700; }}
  td.qty {{ text-align: center; font-weight: 700; color: #003087; }}
  td.status {{ font-weight: 700; color: #E8001D; text-align: center; }}
  .footer {{ margin-top: 20px; font-size: 10px; color: #94A3B8; text-align: center; border-top: 1px solid #eee; padding-top: 10px; }}
  @media print {{
    body {{ padding: 8px; font-size: 10px; }}
    .no-print {{ display: none !important; }}
    @page {{ margin: 10mm; size: A4 landscape; }}
  }}
</style>
</head>
<body>
<div class="header">
  <div class="header-left">
    <h1>🥛 Amul Daily Order Sheet</h1>
    <p>Date: <strong>{date_filter}</strong> &nbsp;|&nbsp; Route: <strong>{route_txt}</strong> &nbsp;|&nbsp; Customers Ordered: <strong>{len(pivot_rows)}</strong> &nbsp;|&nbsp; Not Ordered: <strong>{len(missing)}</strong></p>
  </div>
  <div class="header-right">
    Generated: {now_ist().strftime('%d-%m-%Y %H:%M')} IST<br>
    <button class="no-print" onclick="window.print()" style="margin-top:6px;padding:6px 14px;background:#003087;color:#fff;border:none;border-radius:6px;font-size:11px;font-weight:700;cursor:pointer">🖨️ Print / Save PDF</button>
  </div>
</div>

<div class="section-title">Order Summary (Qty per Customer)</div>
<table>
  <thead>{header_html}</thead>
  <tbody>{data_rows_html}</tbody>
</table>

<div class="footer">
  Amul Wholesale Portal &nbsp;·&nbsp; {date_filter} &nbsp;·&nbsp; {now_ist().strftime('%H:%M')} IST
</div>
</body>
</html>"""
    return Response(html, mimetype="text/html")


if __name__=="__main__":
    load_admin()
    app.run(host="0.0.0.0",port=int(os.environ.get("PORT",5000)),debug=False)
